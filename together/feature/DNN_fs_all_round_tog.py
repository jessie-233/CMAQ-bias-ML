import numpy as np
import matplotlib.pyplot as plt
from sklearn import preprocessing
from tensorflow import keras
from tensorflow.keras import layers
import tensorflow as tf
import tensorflow.keras.backend as K
from tensorflow.keras.callbacks import LearningRateScheduler
import math
from sklearn import metrics
import pandas as pd
from openpyxl import load_workbook

np.set_printoptions(suppress=True) #抑制使用对小数的科学记数法
np.set_printoptions(precision=2) #设置输出的精度

data_BTH = np.load("D:/project/data/BTH/dataset_BTH.npy") 
data_YRD = np.load("D:/project/data/YRD/dataset_YRD.npy") 
data = np.concatenate((data_BTH,data_YRD),axis=0)
dataset_all = data.reshape((-1,43))
train_data = np.concatenate((data[:,0:20,:],data[:,29:49,:],data[:,57:77,:],data[:,88:108,:],data[:,118:138,:],data[:,149:169,:],data[:,179:199,:],data[:,210:230,:],data[:,241:261,:],data[:,271:291,:],data[:,302:322,:],data[:,332:352,:]),axis=1).reshape((-1,43)) #(10560, 43)
test_data = np.concatenate((data[:,20:29,:],data[:,49:57,:],data[:,77:88,:],data[:,108:118,:],data[:,138:149,:],data[:,169:179,:],data[:,199:210,:],data[:,230:241,:],data[:,261:271,:],data[:,291:302,:],data[:,322:332,:],data[:,352:363,:]),axis=1).reshape((-1,43)) #(5412, 43)

var_dict = {'PM2.5_Bias':0, 'PM10_Bias':1, 'NO2_Bias':2, 'SO2_Bias':3, 'O3_Bias':4, 'CO_Bias':5, 'PM2.5_Obs':6, 'PM10_Obs':7, 'NO2_Obs':8, 'SO2_Obs':9, 'O3_Obs':10, 'CO_Obs':11, 'PM2.5_Sim':12, 'PM10_Sim':13, 'NO2_Sim':14, 'SO2_Sim':15, 'O3_Sim':16, 'CO_Sim':17, 'RH_Bias':18, 'TEM_Bias':19, 'WSPD_Bias':20, 'WDIR_Bias':21, 'PRE_Bias':22, 'RH_Obs':23, 'TEM_Obs':24, 'WSPD_Obs':25, 'WDIR_Obs':26, 'PRE_Obs':27, 'PBLH_Sim':28, 'SOLRAD_Sim':29, 'RH_Sim':30, 'TEM_Sim':31, 'WSPD_Sim':32, 'WDIR_Sim':33, 'PRE_Sim':34, 'PM2.5_Bias_ystd':35, 'NO2_Bias_ystd':36, 'RH_Bias_ystd':37, 'O3_Bias_ystd':38, 'SO2_Bias_ystd':39, 'WSPD_Bias_ystd':40, 'NO2_Obs_ystd':41, 'O3_Obs_ystd':42}

early_stopping = keras.callbacks.EarlyStopping(
    monitor='val_loss', 
    patience=10, 
    restore_best_weights=True)
    
#initial_learning_rate / (1 + decay_rate * step / decay_step)
lr_schedule = tf.keras.optimizers.schedules.InverseTimeDecay(
    0.001, 
    decay_steps=160*100, 
    decay_rate=1,
    staircase=False)

def build_model():
    model = keras.Sequential([
    layers.Dense(16, activation='relu', input_shape=[X_train.shape[1]]),
    layers.Dense(4, activation='relu'),
    layers.Dense(4, activation='relu'),
    layers.Dense(3, activation='relu'),
    layers.Dense(1, activation='linear')])
    optimizer=tf.keras.optimizers.RMSprop(lr_schedule)
    model.compile(loss='mse', optimizer=optimizer)
    return model

#prepare x & y dataset
def get_xy_dataset(input_dataset,variables):
    global scaler1,scaler2
    Y = input_dataset[:,0] #'PM2.5_Bias'
    X = np.zeros((len(input_dataset),len(variables)))
    c = 0
    for var in variables:
        X[:,c] = input_dataset[:,var_dict.get(var)]
        c += 1
    X = scaler1.transform(X) #标准化
    Y = Y.reshape((Y.shape[0],1))
    Y = scaler2.transform(Y) #标准化
    return X, Y

# seeds = [111,222,333,444,555,666,777]
seeds = [111,222,333]
#loop for different seeds
for seed in seeds:
    mete_bias = ['RH_Bias','TEM_Bias','WDIR_Bias','WSPD_Bias','PRE_Bias']
    var_bias = ['SO2_Bias','O3_Bias','RH_Bias','TEM_Bias','WDIR_Bias','WSPD_Bias','PRE_Bias']
    #first round of selection
    np.random.seed(seed)
    tf.random.set_seed(seed)
    result_1 = np.zeros((5,1)) #to store result of first round
    np.random.shuffle(train_data)
    row_1 = 0
    for m in mete_bias:
        # NO2_Bias SO2_Bias O3_Bias      RH_Bias TEM_Bias WDIR_Bias WSPD_Bias PRE_Bias
        # NO2_Obs SO2_Obs O3_Obs         RH_Obs  TEM_Obs WSPD_Obs PRE_Obs PBLH_Sim SOLRAD_Sim
        var_sele_1 = ['PM2.5_Bias_ystd','PM2.5_Sim','NO2_Bias',m]
        #standardization: scaler of dataset_all
        Y_all = dataset_all[:,0].reshape((dataset_all.shape[0],1)) #(.,1)
        X_all = np.zeros((len(dataset_all),len(var_sele_1))) #(.,4)
        i = 0
        for var in var_sele_1:
            X_all[:,i] = dataset_all[:,var_dict.get(var)]
            i += 1
        scaler1 = preprocessing.StandardScaler().fit(X_all)
        scaler2 = preprocessing.StandardScaler().fit(Y_all)
        # X_all = scaler1.transform(X_all)
        # Y_all = scaler2.transform(Y_all)
        #train dataset & test dataset
        X_train , y_train = get_xy_dataset(train_data,var_sele_1)
        X_test, y_test = get_xy_dataset(test_data,var_sele_1)
        #build model
        model = build_model()
        model.summary()
        EPOCHS = 500
        history = model.fit(X_train, y_train, epochs=EPOCHS, batch_size=60, validation_split = 0.2, callbacks=[early_stopping], shuffle=False)
        #RMSE on test dataset
        y_pred = model.predict(X_test)
        y_pred = scaler2.inverse_transform(y_pred) #invers
        y_test = scaler2.inverse_transform(y_test) #invers
        test_rmse = math.sqrt(metrics.mean_squared_error(y_test, y_pred))
        print('RMSE on test dataset：%.3f' % test_rmse)
        result_1[row_1,0] = test_rmse
        row_1 += 1
    writer = pd.ExcelWriter('D:/project/data/together/feature/feature_selection_temp.xlsx',engine='openpyxl')
    book = load_workbook('D:/project/data/together/feature/feature_selection_temp.xlsx')
    writer.book = book
    result_1_df = pd.DataFrame(result_1,columns=['NO2_Bias'],index=mete_bias)
    result_1_df.to_excel(writer,sheet_name=str(seed)+'_r1',float_format='%.3f')
    writer.close()
    min_index = str(result_1_df.iloc[:,0].astype("float64").idxmin())
    min_value = result_1_df.loc[min_index,'NO2_Bias']

    #second round of selection
    var_bias.remove(min_index) #remove the chosen mete_bias var
    var_base = ['PM2.5_Bias_ystd','PM2.5_Sim','NO2_Bias',min_index]
    def Rounding2(vars,benchmark):
        global var_base,dataset_all,var_dict,scaler1,scaler2,var_sele_3,X_train
        result = pd.DataFrame(columns=['RMSE_test','Improvement'],index=range(len(vars)+1))
        row_2 = 0 # row number
        result.iloc[row_2,0] = benchmark
        for j in vars: #choose one addition var every time
            row_2 += 1
            var_sele_3 = var_base + [j]
            Y_all = dataset_all[:,0].reshape((dataset_all.shape[0],1))
            X_all = np.zeros((len(dataset_all),len(var_sele_3)))
            i = 0
            for k in var_sele_3: #make x dataset
                X_all[:,i] = dataset_all[:,var_dict.get(k)]
                i += 1
            scaler1 = preprocessing.StandardScaler().fit(X_all)
            scaler2 = preprocessing.StandardScaler().fit(Y_all)
            # X_all = scaler1.transform(X_all)
            # Y_all = scaler2.transform(Y_all)
            #x,y dataset
            X_train , y_train = get_xy_dataset(train_data,var_sele_3)
            X_test, y_test = get_xy_dataset(test_data,var_sele_3)
            #build model
            model = build_model()
            model.summary()
            EPOCHS = 500
            model.fit(X_train, y_train, epochs=EPOCHS, batch_size=60, validation_split = 0.2, callbacks=[early_stopping], shuffle=False)
            #RMSE on test dataset
            y_pred = model.predict(X_test)
            y_pred = scaler2.inverse_transform(y_pred) #invers
            y_test = scaler2.inverse_transform(y_test) #invers
            test_rmse = math.sqrt(metrics.mean_squared_error(y_test, y_pred))
            impro = benchmark - test_rmse
            print('RMSE on test dataset：%.3f' % test_rmse)
            result.iloc[row_2,0] = test_rmse
            result.iloc[row_2,1] = impro
        result.index=['benchmark']+vars
        return result

    benchmark = min_value
    result_2 = pd.DataFrame(columns=['RMSE_test','Improvement'])
    while True:
        rounding_result_2 = Rounding2(var_bias,benchmark) #dataframe (7,2)
        result_2 = result_2.append(rounding_result_2)
        var_bias = [str(a) for a in rounding_result_2.index[1:]] #Index dtype=object
        max_index = str(rounding_result_2.iloc[1:,1].astype("float64").idxmax())
        max_value = rounding_result_2.loc[max_index,'Improvement'] #float
        if max_value >= 0.2:
            benchmark = rounding_result_2.loc[max_index,'RMSE_test'] #float
            var_bias.remove(max_index)
            var_base = var_base + [max_index]
        else:    
            break
    writer = pd.ExcelWriter('D:/project/data/together/feature/feature_selection_temp.xlsx',engine='openpyxl')
    book = load_workbook('D:/project/data/together/feature/feature_selection_temp.xlsx')
    writer.book = book
    result_2.to_excel(writer, sheet_name=str(seed)+'_r2',float_format='%.3f')
    writer.save()
    writer.close()

    #third round of selection
    var_obs = ['SO2_Obs','O3_Obs','RH_Obs','TEM_Obs','WSPD_Obs','PRE_Obs','PBLH_Sim','SOLRAD_Sim'] #删除了NO2_Obs
    def Rounding3(vars,benchmark):
        global var_base,dataset_all,var_dict,scaler1,scaler2,var_sele_3,X_train
        result = pd.DataFrame(columns=['RMSE_test','Improvement'],index=range(len(vars)+1))
        row_3 = 0 # row number
        result.iloc[row_3,0] = benchmark
        for j in vars: #choose one addition var every time
            row_3 += 1
            var_sele_3 = var_base + [j]
            Y_all = dataset_all[:,0].reshape((dataset_all.shape[0],1))
            X_all = np.zeros((len(dataset_all),len(var_sele_3)))
            i = 0
            for k in var_sele_3: #make x dataset
                X_all[:,i] = dataset_all[:,var_dict.get(k)]
                i += 1
            scaler1 = preprocessing.StandardScaler().fit(X_all)
            scaler2 = preprocessing.StandardScaler().fit(Y_all)
            # X_all = scaler1.transform(X_all)
            # Y_all = scaler2.transform(Y_all)
            #x,y dataset
            X_train , y_train = get_xy_dataset(train_data,var_sele_3)
            X_test, y_test = get_xy_dataset(test_data,var_sele_3)
            #build model
            model = build_model()
            model.summary()
            EPOCHS = 500
            model.fit(X_train, y_train, epochs=EPOCHS, batch_size=60, validation_split = 0.2, callbacks=[early_stopping], shuffle=False)
            #RMSE on test dataset
            y_pred = model.predict(X_test)
            y_pred = scaler2.inverse_transform(y_pred) #invers
            y_test = scaler2.inverse_transform(y_test) #invers
            test_rmse = math.sqrt(metrics.mean_squared_error(y_test, y_pred))
            impro = benchmark - test_rmse
            print('RMSE on test dataset：%.3f' % test_rmse)
            result.iloc[row_3,0] = test_rmse
            result.iloc[row_3,1] = impro
        result.index=['benchmark']+vars
        return result

    result_3 = pd.DataFrame(columns=['RMSE_test','Improvement'])
    while True:
        if 'SO2_Bias' in var_base and 'SO2_Obs' in var_obs:
            var_obs.remove('SO2_Obs')
        if 'O3_Obs' in var_base and 'TEM_Obs' in var_obs:
            var_obs.remove('TEM_Obs')
        if 'O3_Obs' in var_base and 'SOLRAD_Sim' in var_obs:
            var_obs.remove('SOLRAD_Sim')
        if 'TEM_Obs' in var_base and 'SOLRAD_Sim' in var_obs:
            var_obs.remove('SOLRAD_Sim')
        rounding_result_3 = Rounding3(var_obs,benchmark)
        result_3 = result_3.append(rounding_result_3)
        var_obs = [str(a) for a in rounding_result_3.index[1:]]
        max_index = str(rounding_result_3.iloc[1:,1].astype("float64").idxmax())
        max_value = rounding_result_3.loc[max_index,'Improvement']
        if max_value >= 0.2:
            benchmark = rounding_result_3.loc[max_index,'RMSE_test']
            var_obs.remove(max_index)
            var_base = var_base + [max_index]
        else:    
            break
    writer = pd.ExcelWriter('D:/project/data/together/feature/feature_selection_temp.xlsx',engine='openpyxl')
    book = load_workbook('D:/project/data/together/feature/feature_selection_temp.xlsx')
    writer.book = book
    result_3.to_excel(writer, sheet_name=str(seed)+'_r3',float_format='%.3f')
    writer.save()
    writer.close()
    print(var_base)




    


